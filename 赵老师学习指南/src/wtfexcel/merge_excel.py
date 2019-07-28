from openpyxl import load_workbook
import os
filepath = os.getcwd() + '/origin/'

def add(dst_sheet, src_sheet):
    for row in src_sheet.rows:
        dst_sheet.append([cell.value for cell in list(row)])

wb = load_workbook(filename = filepath + 'merge.xlsx')
sheet1 = wb['Sheet1']
sheet2 = wb['Sheet2']
merged_sheet = wb.create_sheet('merged')

add(merged_sheet, sheet1)
add(merged_sheet, sheet2)

wb.save(filepath + 'merge.xlsx')

