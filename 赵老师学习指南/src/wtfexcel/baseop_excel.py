from openpyxl import load_workbook
from openpyxl import Workbook
import os
filepath = os.getcwd() + '/origin/'

# 一般行列转换成excel行列 比如(1, 1)=>A1
# 别看这个了 我好无聊
def indexChange(row, column):
    if not isinstance(row, int) or not isinstance(column, int):
        return
    indexRes = ""    
    while column > 25:
        indexRes = indexRes + chr(column % 26 + ord('A') - 1)
        column = int(column / 26)
    indexRes = indexRes + chr(column % 26 + ord('A') - 1)
    return indexRes[::-1] + str(row)

# 读取sample.xlsx
def read_demo():
    # 相对路径就可以了 windows直接去掉filepath
    wb = load_workbook(filename = filepath + 'sample.xlsx')
    sheet = wb.active

    # 直接返回行 列对象
    for row in sheet.rows:
        for cell in row:
            print(cell.value)

    # 根据行列号去读
    for i in range(1, sheet.max_row + 1):
        for j in range(1, sheet.max_column + 1):
            print(sheet.cell(row = i, column = j).value)


def write_demo():
    wb = Workbook()
    filename = 'sample.xlsx'
    sheet = wb.active
    # 10行10列
    for i in range(1, 5):
        for j in range(1, 5):
            index = indexChange(i, j)
            sheet[index] = index
    wb.save(filepath + filename)

# 如果你跑这个.py文件的话 就会被执行 可以理解为主函数吧
if __name__ == '__main__':
    # write_demo()
    read_demo()

