from openpyxl import load_workbook
import os
filepath = os.getcwd()

wb = load_workbook(filename = filepath + '/sample.xlsx')

# current sheet
sheet = wb.active
print(sheet.cell(row = 1, column = 1).value)
print(sheet.cell(row = 1, column = 2).value)